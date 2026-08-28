from __future__ import annotations

import csv
import io
import math
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree

import openpyxl
from charset_normalizer import from_bytes
from converted_document import ConversionResult, DocumentConversionError

from converters.base import (
    ConversionLimits,
    DocumentConverter,
    LocatedTextBlock,
)
from converters.plaintext import _detect_encoding, _looks_binary


@dataclass(frozen=True)
class OoxmlZipLimits:
    max_entries: int = 5_000
    max_uncompressed_bytes: int = 128 * 1024 * 1024
    max_entry_bytes: int = 32 * 1024 * 1024
    max_compression_ratio: float = 200.0

    def __post_init__(self) -> None:
        if (
            self.max_entries <= 0
            or self.max_uncompressed_bytes <= 0
            or self.max_entry_bytes <= 0
            or self.max_compression_ratio <= 0
        ):
            raise ValueError("OOXML ZIP limits must be positive")


@dataclass(frozen=True)
class SpreadsheetLimits:
    max_sheets: int = 100
    max_rows_per_sheet: int = 20_000
    max_columns: int = 256
    max_cell_chars: int = 65_536
    rows_per_unit: int = 50

    def __post_init__(self) -> None:
        if (
            self.max_sheets <= 0
            or self.max_rows_per_sheet <= 0
            or self.max_columns <= 0
            or self.max_cell_chars <= 0
            or self.rows_per_unit <= 0
        ):
            raise ValueError("spreadsheet limits must be positive")


class XlsxConverter(DocumentConverter):
    document_kind = "xlsx"
    conversion_profile = "xlsx-v1"
    converter_name = "xlsx-local"
    source_type = "xlsx"

    def __init__(
        self,
        limits: ConversionLimits | None = None,
        *,
        spreadsheet_limits: SpreadsheetLimits | None = None,
        zip_limits: OoxmlZipLimits | None = None,
    ):
        super().__init__(limits)
        self.spreadsheet_limits = spreadsheet_limits or SpreadsheetLimits()
        self.zip_limits = zip_limits or OoxmlZipLimits()

    def convert(
        self, source_file: str | Path, output_dir: str | Path
    ) -> ConversionResult:
        inspection = self._inspect_source(source_file)
        if inspection.path.suffix.lower() in {".xlsm", ".xltm", ".xlam"}:
            raise DocumentConversionError(
                "MACRO_ENABLED_DOCUMENT_UNSUPPORTED",
                "macro-enabled spreadsheet formats are not supported",
            )

        package = inspect_ooxml_package(
            inspection.path,
            self.zip_limits,
            required_entries=("[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"),
        )
        if package.has_macros:
            raise DocumentConversionError(
                "MACRO_ENABLED_DOCUMENT_UNSUPPORTED",
                "macro-enabled spreadsheet packages are not supported",
            )

        warnings: list[str] = []
        formula_without_cache = _count_formulas_without_cached_values(
            inspection.path, package.names
        )
        if formula_without_cache:
            warnings.append(
                f"XLSX_FORMULAS_WITHOUT_CACHED_VALUES:{formula_without_cache}"
            )
        if any(
            name.casefold().startswith("xl/externallinks/") for name in package.names
        ):
            warnings.append("XLSX_EXTERNAL_LINKS_IGNORED")

        try:
            workbook = openpyxl.load_workbook(
                inspection.path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except (OSError, ValueError, KeyError, zipfile.BadZipFile) as error:
            raise DocumentConversionError(
                "XLSX_PARSE_FAILED", "spreadsheet package could not be parsed"
            ) from error

        try:
            if len(workbook.worksheets) > self.spreadsheet_limits.max_sheets:
                raise DocumentConversionError(
                    "XLSX_TOO_MANY_SHEETS",
                    "spreadsheet exceeds the configured sheet limit",
                )

            units: list[LocatedTextBlock] = []
            has_cells = False
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                max_row = worksheet.max_row or 0
                max_column = worksheet.max_column or 0
                if max_row > self.spreadsheet_limits.max_rows_per_sheet:
                    raise DocumentConversionError(
                        "XLSX_TOO_MANY_ROWS",
                        "spreadsheet exceeds the configured row limit",
                    )
                if max_column > self.spreadsheet_limits.max_columns:
                    raise DocumentConversionError(
                        "XLSX_TOO_MANY_COLUMNS",
                        "spreadsheet exceeds the configured column limit",
                    )

                sheet_name = _normalize_text(
                    worksheet.title, self.spreadsheet_limits.max_cell_chars
                )
                units.append(
                    LocatedTextBlock(
                        f"# Sheet: {_escape_heading(sheet_name)}\n",
                        {
                            "type": "xlsx",
                            "kind": "sheet",
                            "sheet": sheet_name,
                            "sheet_index": sheet_index,
                        },
                    )
                )

                block: list[tuple[int, list[str]]] = []
                for row_number, row in enumerate(worksheet.iter_rows(), start=1):
                    if row_number > self.spreadsheet_limits.max_rows_per_sheet:
                        raise DocumentConversionError(
                            "XLSX_TOO_MANY_ROWS",
                            "spreadsheet exceeds the configured row limit",
                        )
                    values = [
                        _format_cell(cell.value, self.spreadsheet_limits.max_cell_chars)
                        for cell in row
                    ]
                    while values and values[-1] == "":
                        values.pop()
                    if len(values) > self.spreadsheet_limits.max_columns:
                        raise DocumentConversionError(
                            "XLSX_TOO_MANY_COLUMNS",
                            "spreadsheet exceeds the configured column limit",
                        )
                    if any(values):
                        has_cells = True
                    block.append((row_number, values))
                    if len(block) == self.spreadsheet_limits.rows_per_unit:
                        _append_xlsx_block(units, block, sheet_name, sheet_index)
                        block = []
                if block:
                    _append_xlsx_block(units, block, sheet_name, sheet_index)

            if not has_cells:
                raise DocumentConversionError(
                    "EMPTY_DOCUMENT", "spreadsheet has no embeddable cell values"
                )

            return self._convert_located_blocks(
                inspection,
                output_dir,
                units,
                "binary",
                tuple(warnings),
            )
        except DocumentConversionError:
            raise
        except (OSError, ValueError, TypeError, ElementTree.ParseError) as error:
            raise DocumentConversionError(
                "XLSX_PARSE_FAILED", "spreadsheet package could not be parsed"
            ) from error
        finally:
            workbook.close()


class CsvConverter(DocumentConverter):
    document_kind = "csv"
    conversion_profile = "csv-v1"
    converter_name = "csv-local"
    source_type = "csv"

    def __init__(
        self,
        limits: ConversionLimits | None = None,
        *,
        spreadsheet_limits: SpreadsheetLimits | None = None,
    ):
        super().__init__(limits)
        self.spreadsheet_limits = spreadsheet_limits or SpreadsheetLimits()

    def convert(
        self, source_file: str | Path, output_dir: str | Path
    ) -> ConversionResult:
        inspection = self._inspect_source(source_file)
        data = self._read_bounded_bytes(inspection)
        encoding = _detect_csv_encoding(data)
        try:
            text = data.decode(encoding, errors="strict")
        except UnicodeDecodeError as error:
            raise DocumentConversionError(
                "CSV_DECODE_FAILED", "CSV source could not be decoded"
            ) from error
        if _looks_binary(text):
            raise DocumentConversionError(
                "CSV_BINARY_CONTENT", "CSV source contains binary content"
            )

        delimiter = _detect_csv_delimiter(text)
        units: list[LocatedTextBlock] = []
        block: list[tuple[int, list[str]]] = []
        has_cells = False
        try:
            reader = csv.reader(
                io.StringIO(text, newline=""), delimiter=delimiter, strict=True
            )
            for row_number, row in enumerate(reader, start=1):
                if row_number > self.spreadsheet_limits.max_rows_per_sheet:
                    raise DocumentConversionError(
                        "CSV_TOO_MANY_ROWS",
                        "CSV source exceeds the configured row limit",
                    )
                if len(row) > self.spreadsheet_limits.max_columns:
                    raise DocumentConversionError(
                        "CSV_TOO_MANY_COLUMNS",
                        "CSV source exceeds the configured column limit",
                    )
                values = [
                    _normalize_text(value, self.spreadsheet_limits.max_cell_chars)
                    for value in row
                ]
                while values and values[-1] == "":
                    values.pop()
                has_cells = has_cells or any(values)
                block.append((row_number, values))
                if len(block) == self.spreadsheet_limits.rows_per_unit:
                    _append_csv_block(units, block)
                    block = []
        except csv.Error as error:
            raise DocumentConversionError(
                "CSV_PARSE_FAILED", "CSV source could not be parsed"
            ) from error
        if block:
            _append_csv_block(units, block)
        if not has_cells:
            raise DocumentConversionError(
                "EMPTY_DOCUMENT", "CSV source has no embeddable cell values"
            )

        # A non-comma delimiter is not a conversion defect: the file parsed
        # cleanly. Reporting it as a warning made every single CSV land in
        # completed_with_warnings, which drained the status of meaning and hid
        # the documents that really did convert imperfectly.
        notes = (f"CSV_DELIMITER:{_delimiter_name(delimiter)}",)
        return self._convert_located_blocks(
            inspection,
            output_dir,
            units,
            encoding,
            (),
            notes,
        )


@dataclass(frozen=True)
class OoxmlPackageInspection:
    names: tuple[str, ...]
    has_macros: bool


def inspect_ooxml_package(
    path: Path,
    limits: OoxmlZipLimits,
    *,
    required_entries: Sequence[str],
) -> OoxmlPackageInspection:
    try:
        with zipfile.ZipFile(path) as archive:
            entries = [entry for entry in archive.infolist() if not entry.is_dir()]
            if len(entries) > limits.max_entries:
                raise DocumentConversionError(
                    "OOXML_ZIP_TOO_MANY_ENTRIES",
                    "OOXML package exceeds the configured entry limit",
                )

            names: list[str] = []
            seen: set[str] = set()
            total_uncompressed = 0
            total_compressed = 0
            for entry in entries:
                name = entry.filename
                normalized = PurePosixPath(name.replace("\\", "/"))
                if (
                    normalized.is_absolute()
                    or ".." in normalized.parts
                    or not normalized.parts
                ):
                    raise DocumentConversionError(
                        "OOXML_UNSAFE_ZIP_ENTRY",
                        "OOXML package contains an unsafe entry path",
                    )
                canonical = normalized.as_posix()
                folded = canonical.casefold()
                if folded in seen:
                    raise DocumentConversionError(
                        "OOXML_DUPLICATE_ZIP_ENTRY",
                        "OOXML package contains duplicate entry names",
                    )
                seen.add(folded)
                names.append(canonical)
                if entry.flag_bits & 0x1:
                    raise DocumentConversionError(
                        "OOXML_ENCRYPTED_ENTRY",
                        "encrypted OOXML package entries are not supported",
                    )
                if entry.file_size > limits.max_entry_bytes:
                    raise DocumentConversionError(
                        "OOXML_ZIP_ENTRY_TOO_LARGE",
                        "OOXML package entry exceeds the configured size limit",
                    )
                total_uncompressed += entry.file_size
                total_compressed += entry.compress_size
                if total_uncompressed > limits.max_uncompressed_bytes:
                    raise DocumentConversionError(
                        "OOXML_ZIP_UNCOMPRESSED_TOO_LARGE",
                        "OOXML package exceeds the configured expanded size limit",
                    )
                ratio = entry.file_size / max(entry.compress_size, 1)
                if ratio > limits.max_compression_ratio:
                    raise DocumentConversionError(
                        "OOXML_ZIP_COMPRESSION_RATIO_EXCEEDED",
                        "OOXML package entry exceeds the configured compression ratio",
                    )

            if (
                total_uncompressed / max(total_compressed, 1)
                > limits.max_compression_ratio
            ):
                raise DocumentConversionError(
                    "OOXML_ZIP_COMPRESSION_RATIO_EXCEEDED",
                    "OOXML package exceeds the configured compression ratio",
                )
            name_set = set(names)
            if any(required not in name_set for required in required_entries):
                raise DocumentConversionError(
                    "OOXML_INVALID_STRUCTURE",
                    "OOXML package is missing required document parts",
                )

            for entry in entries:
                folded_name = entry.filename.casefold()
                if folded_name.endswith((".xml", ".rels")):
                    _reject_unsafe_xml(archive, entry)

            bad_entry = archive.testzip()
            if bad_entry is not None:
                raise DocumentConversionError(
                    "OOXML_CORRUPT_ZIP_ENTRY",
                    "OOXML package contains a corrupt ZIP entry",
                )
            content_types = archive.read("[Content_Types].xml")
            lowered_content_types = content_types.lower()
            has_macros = any(
                name.casefold().endswith("vbaproject.bin") for name in names
            ) or any(
                marker in lowered_content_types
                for marker in (b"macroenabled", b"application/vnd.ms-office.vbaproject")
            )
            return OoxmlPackageInspection(tuple(names), has_macros)
    except DocumentConversionError:
        raise
    except (zipfile.BadZipFile, zipfile.LargeZipFile, OSError, RuntimeError) as error:
        raise DocumentConversionError(
            "OOXML_INVALID_ZIP", "source is not a valid OOXML ZIP package"
        ) from error


def _reject_unsafe_xml(archive: zipfile.ZipFile, entry: zipfile.ZipInfo) -> None:
    tail = b""
    try:
        with archive.open(entry) as stream:
            while chunk := stream.read(1024 * 1024):
                probe = (tail + chunk).lower().replace(b"\x00", b"")
                if b"<!doctype" in probe or b"<!entity" in probe:
                    raise DocumentConversionError(
                        "OOXML_UNSAFE_XML",
                        "OOXML package contains a prohibited XML declaration",
                    )
                tail = probe[-32:]
    except DocumentConversionError:
        raise
    except (OSError, RuntimeError, zipfile.BadZipFile) as error:
        raise DocumentConversionError(
            "OOXML_CORRUPT_ZIP_ENTRY", "OOXML package contains an unreadable XML entry"
        ) from error


def _append_xlsx_block(
    units: list[LocatedTextBlock],
    block: list[tuple[int, list[str]]],
    sheet_name: str,
    sheet_index: int,
) -> None:
    nonempty = [(row_number, values) for row_number, values in block if any(values)]
    if not nonempty:
        return
    row_start = nonempty[0][0]
    row_end = nonempty[-1][0]
    units.append(
        LocatedTextBlock(
            _markdown_row_table(nonempty),
            {
                "type": "xlsx",
                "kind": "table",
                "sheet": sheet_name,
                "sheet_index": sheet_index,
                "row_start": row_start,
                "row_end": row_end,
            },
        )
    )


def _append_csv_block(
    units: list[LocatedTextBlock], block: list[tuple[int, list[str]]]
) -> None:
    nonempty = [(row_number, values) for row_number, values in block if any(values)]
    if not nonempty:
        return
    units.append(
        LocatedTextBlock(
            _markdown_row_table(nonempty),
            {
                "type": "csv",
                "kind": "table",
                "row_start": nonempty[0][0],
                "row_end": nonempty[-1][0],
            },
        )
    )


def _markdown_row_table(rows: list[tuple[int, list[str]]]) -> str:
    width = max((len(values) for _, values in rows), default=0)
    headers = ["Row", *(_column_name(index) for index in range(1, width + 1))]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row_number, values in rows:
        padded = [*values, *("" for _ in range(width - len(values)))]
        lines.append(
            "| "
            + " | ".join(
                [str(row_number), *(_escape_table_cell(value) for value in padded)]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _detect_csv_delimiter(text: str) -> str:
    sample = text[:65_536]
    if not sample.strip():
        return ","
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        candidates = [delimiter for delimiter in ",;\t|" if delimiter in sample]
        if not candidates:
            return ","
        raise DocumentConversionError(
            "CSV_DELIMITER_UNDETECTABLE",
            "CSV delimiter could not be detected reliably",
        )


def _detect_csv_encoding(data: bytes) -> str:
    detected = _detect_encoding(data)
    if detected.casefold() in {"utf-8", "utf_8", "utf-8-sig", "utf-16", "utf-32"}:
        return detected

    east_asian_matches: list[tuple[float, int, str]] = []
    for order, match in enumerate(from_bytes(data)):
        encoding = match.encoding.casefold().replace("-", "_")
        family = _east_asian_encoding_family(encoding)
        if family is None or match.percent_chaos > 20:
            continue
        decoded = str(match)
        han = sum(1 for char in decoded if "\u3400" <= char <= "\u9fff")
        hangul = sum(1 for char in decoded if "\uac00" <= char <= "\ud7af")
        kana = sum(1 for char in decoded if "\u3040" <= char <= "\u30ff")
        east_asian_chars = max(han + hangul + kana, 1)
        if family == "chinese":
            script_score = 2 * han - 3 * (hangul + kana) + 20 * han / east_asian_chars
        elif family == "korean":
            script_score = 3 * hangul + han - 3 * kana + 20 * hangul / east_asian_chars
        else:
            script_score = 3 * kana + han - 3 * hangul + 20 * kana / east_asian_chars
        score = script_score - match.percent_chaos
        east_asian_matches.append((score, -order, match.encoding))
    if east_asian_matches:
        best_score, _order, best_encoding = max(east_asian_matches)
        if best_score > 0:
            return best_encoding
    return detected


def _east_asian_encoding_family(encoding: str) -> str | None:
    if encoding.startswith(("gb", "big5", "hz")):
        return "chinese"
    if encoding.startswith(("cp949", "euc_kr", "johab")):
        return "korean"
    if encoding.startswith(("shift_jis", "cp932", "euc_j", "iso2022_jp")):
        return "japanese"
    return None


def _delimiter_name(delimiter: str) -> str:
    return {",": "comma", ";": "semicolon", "\t": "tab", "|": "pipe"}[delimiter]


def _format_cell(value: Any, max_chars: int) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        if math.isinf(value):
            return "Infinity" if value > 0 else "-Infinity"
    return _normalize_text(str(value), max_chars)


def _normalize_text(value: str, max_chars: int) -> str:
    normalized = value.replace("\r\n", "\n").replace("\r", "\n")
    if len(normalized) > max_chars:
        raise DocumentConversionError(
            "SPREADSHEET_CELL_TOO_LARGE",
            "spreadsheet cell exceeds the configured character limit",
        )
    return normalized


def _escape_heading(value: str) -> str:
    return value.replace("\n", " ").replace("#", "\\#")


def _escape_table_cell(value: str) -> str:
    return value.replace("\\", "\\\\").replace("|", "\\|").replace("\n", "<br>")


def _column_name(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _count_formulas_without_cached_values(path: Path, names: Sequence[str]) -> int:
    worksheet_names = sorted(
        name
        for name in names
        if name.casefold().startswith("xl/worksheets/")
        and name.casefold().endswith(".xml")
    )
    count = 0
    try:
        with zipfile.ZipFile(path) as archive:
            for name in worksheet_names:
                with archive.open(name) as stream:
                    for _event, element in ElementTree.iterparse(
                        stream, events=("end",)
                    ):
                        if element.tag.rsplit("}", 1)[-1] != "c":
                            continue
                        formula = None
                        cached = None
                        for child in element:
                            local_name = child.tag.rsplit("}", 1)[-1]
                            if local_name == "f":
                                formula = child
                            elif local_name == "v":
                                cached = child
                        if formula is not None and (
                            cached is None or cached.text in (None, "")
                        ):
                            count += 1
                        element.clear()
    except (OSError, zipfile.BadZipFile, ElementTree.ParseError) as error:
        raise DocumentConversionError(
            "XLSX_PARSE_FAILED", "worksheet XML could not be parsed"
        ) from error
    return count
